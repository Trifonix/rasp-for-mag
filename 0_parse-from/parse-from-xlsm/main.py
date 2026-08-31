"""
Парсер расписания из Excel (.xlsm / .xlsx) для группы 12-25РПм.

Использование:
  python 0_parse-from/parse-from-xlsm/main.py
  python 0_parse-from/parse-from-xlsm/main.py путь/к/файлу.xlsm
  python 0_parse-from/parse-from-xlsm/main.py --merge schedule3.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "0_parse-from" / "2026-08-31" / "2026-08-31.xlsm"
DEFAULT_OUTPUT = ROOT / "schedule3.json"
DEFAULT_GROUP = "12-25РПм"
HEADER_ROW = 10

WEEKDAYS_MAP = {
    0: "понедельник",
    1: "вторник",
    2: "среда",
    3: "четверг",
    4: "пятница",
    5: "суббота",
    6: "воскресенье",
}

PAIR_TIMES = {
    "1": "09:00-10:30",
    "2": "10:40-12:10",
    "3": "12:30-14:00",
    "4": "14:10-15:40",
    "5": "15:50-17:20",
    "6": "17:40-19:10",
    "7": "19:20-20:50",
}

SINGLE_WEEK_RENAME = {
    "Дни недели": "W1_День",
    "пара": "W1_Пара",
    "Вид занятий": "W1_Вид_занятий",
    "Дисциплина": "W1_Дисциплина",
    "Преподаватель": "W1_Преподаватель",
    "Unnamed: 6": "W1_Ученая_степень",
    "Ссылка": "W1_Ссылка",
}

DUAL_WEEK_RENAME = {
    "Дни недели": "W1_День",
    "пара": "W1_Пара",
    "Вид занятий": "W1_Вид_занятий",
    "Дисциплина": "W1_Дисциплина",
    "Преподаватель": "W1_Преподаватель",
    "Unnamed: 5": "W1_Ученая_степень",
    "Ссылка": "W1_Ссылка",
    "Дни недели.1": "W2_День",
    "пара.1": "W2_Пара",
    "Вид занятий.1": "W2_Вид_занятий",
    "Дисциплина.1": "W2_Дисциплина",
    "Преподаватель.1": "W2_Преподаватель",
    "Unnamed: 12": "W2_Ученая_степень",
    "Ссылка.1": "W2_Ссылка",
}


def clean(value) -> str:
    if pd.notna(value):
        text = str(value).replace("\n", " ").strip()
        if text in ("0", "0.0", "0.00", "nan"):
            return ""
        return text
    return ""


def fix_pair(pair_text: str) -> str:
    pair_text = clean(pair_text)
    if not pair_text:
        return ""

    match = re.search(r"(\d+)", pair_text)
    if not match:
        return pair_text.strip()

    num = match.group(1)
    time_str = PAIR_TIMES.get(num, "")
    time_in_text = re.search(r"(\d{1,2})[.:](\d{2})\s*-\s*(\d{1,2})[.:](\d{2})", pair_text)
    if time_in_text:
        start_h, start_m, end_h, end_m = time_in_text.groups()
        time_str = f"{int(start_h):02d}:{start_m}-{int(end_h):02d}:{end_m}"

    return f"{num} пара {time_str}".strip()


def join_teacher(record, teacher_key: str, degree_key: str) -> str:
    teacher = clean(record.get(teacher_key))
    if not teacher:
        return ""
    return teacher


def normalize_day(day_text) -> str | None:
    text = clean(day_text)
    if not text:
        return None

    parts = text.split()
    if len(parts) >= 2 and re.match(r"\d{2}\.\d{2}\.\d{4}", parts[0]):
        return f"{parts[0]} {parts[1].strip()}"

    if re.match(r"\d{4}-\d{2}-\d{2}", text):
        try:
            date_part = text.split()[0]
            dt_obj = pd.to_datetime(date_part)
            weekday_name = WEEKDAYS_MAP.get(dt_obj.weekday(), "")
            return f"{dt_obj.strftime('%d.%m.%Y')} {weekday_name}"
        except Exception:
            return None

    if len(parts) == 1 and parts[0].lower() in WEEKDAYS_MAP.values():
        return None

    return text


def parse_week_block(df: pd.DataFrame, prefix: str) -> dict[str, list[dict]]:
    schedule: dict[str, list[dict]] = {}
    current_day = None

    day_col = f"{prefix}_День"
    pair_col = f"{prefix}_Пара"
    activity_col = f"{prefix}_Вид_занятий"
    discipline_col = f"{prefix}_Дисциплина"
    teacher_col = f"{prefix}_Преподаватель"
    degree_col = f"{prefix}_Ученая_степень"
    link_col = f"{prefix}_Ссылка"

    for _, row in df.iterrows():
        day_raw = row.get(day_col)
        if pd.notna(day_raw):
            normalized = normalize_day(day_raw)
            if normalized:
                current_day = normalized

        discipline = clean(row.get(discipline_col))
        activity = clean(row.get(activity_col))
        pair_raw = row.get(pair_col)

        if not (discipline and activity and clean(pair_raw) and current_day):
            continue

        entry = {
            "Пара": fix_pair(pair_raw),
            "Вид занятий": activity,
            "Дисциплина": discipline,
            "Преподаватель": join_teacher(row, teacher_col, degree_col),
            "Ссылка": clean(row.get(link_col)),
        }
        schedule.setdefault(current_day, []).append(entry)

    return schedule


def load_sheet(file_path: Path, group: str) -> pd.DataFrame:
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if group not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        workbook.close()
        raise ValueError(f"Лист «{group}» не найден. Доступные листы: {available}")
    workbook.close()

    return pd.read_excel(
        file_path,
        sheet_name=group,
        header=HEADER_ROW,
        converters={"Дни недели": str, "Дни недели.1": str},
    )


def parse_schedule(file_path: Path, group: str) -> dict[str, list[dict]]:
    df = load_sheet(file_path, group)
    df = df.dropna(how="all")
    df.columns = [clean(column) for column in df.columns]

    rename_map = DUAL_WEEK_RENAME if "Дни недели.1" in df.columns else SINGLE_WEEK_RENAME
    df.rename(columns=rename_map, inplace=True)

    schedule = parse_week_block(df, "W1")
    if "W2_День" in df.columns:
        schedule_w2 = parse_week_block(df, "W2")
        for day, lessons in schedule_w2.items():
            schedule.setdefault(day, []).extend(lessons)

    return schedule


def sort_schedule(schedule: dict[str, list[dict]]) -> dict[str, list[dict]]:
    def sort_key(date_str: str) -> datetime:
        date_part = date_str.split()[0]
        try:
            return datetime.strptime(date_part, "%d.%m.%Y")
        except ValueError:
            return datetime.max

    sorted_keys = sorted(
        [key for key in schedule if sort_key(key) != datetime.max],
        key=sort_key,
    )
    return {key: schedule[key] for key in sorted_keys}


def merge_schedules(existing: dict[str, list[dict]], new_data: dict[str, list[dict]]) -> dict[str, list[dict]]:
    merged = dict(existing)
    for day, lessons in new_data.items():
        merged[day] = lessons
    return sort_schedule(merged)


def load_json(path: Path) -> dict[str, list[dict]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as file:
        return json.load(file)


def save_json(path: Path, schedule: dict[str, list[dict]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        json.dump(sort_schedule(schedule), file, ensure_ascii=False, indent=2)


def normalize_group(name: str) -> str:
    return name.lower().replace(" ", "")


def resolve_group_name(file_path: Path, group: str) -> str:
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()

    target = normalize_group(group)
    for name in sheet_names:
        if normalize_group(name) == target:
            return name
        if "12-25" in name.lower() and "рп" in name.lower():
            return name

    raise ValueError(f"Не удалось найти лист группы «{group}». Листы: {', '.join(sheet_names)}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Парсинг расписания группы 12-25РПм из Excel")
    parser.add_argument(
        "input",
        nargs="?",
        default=str(DEFAULT_INPUT),
        help="Путь к .xlsm/.xlsx (по умолчанию: Расписание_ИИТ_1_неделя.xlsm)",
    )
    parser.add_argument(
        "--group",
        default=DEFAULT_GROUP,
        help="Имя листа группы (по умолчанию: 12-25РПм)",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Куда сохранить JSON (по умолчанию: schedule3.json в корне проекта)",
    )
    parser.add_argument(
        "--merge",
        action="store_true",
        help="Дополнить существующий JSON новыми днями (дни из файла перезаписываются)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только показать результат, не сохранять файл",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Ошибка: файл не найден: {input_path}", file=sys.stderr)
        return 1

    try:
        group = resolve_group_name(input_path, args.group)
        parsed = parse_schedule(input_path, group)
    except Exception as exc:
        print(f"Ошибка парсинга: {exc}", file=sys.stderr)
        return 1

    if not parsed:
        print("Предупреждение: в файле не найдено занятий для группы.", file=sys.stderr)

    if args.merge:
        schedule = merge_schedules(load_json(output_path), parsed)
    else:
        schedule = sort_schedule(parsed)

    print(f"Лист: {group}")
    print(f"Новых/обновлённых дней: {len(parsed)}")
    for day, lessons in parsed.items():
        print(f"  {day}: {len(lessons)} занятий")

    if args.dry_run:
        print(json.dumps(schedule, ensure_ascii=False, indent=2))
        return 0

    save_json(output_path, schedule)
    print(f"OK: Расписание сохранено: {output_path} ({len(schedule)} дней всего)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
